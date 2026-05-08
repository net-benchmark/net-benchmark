import importlib.util
import json
import os
import tempfile

import pandas as pd
import pytest
from openpyxl import Workbook

import net_benchmark.dns_benchmark.exporters as exporters
from net_benchmark.dns_benchmark.analysis import BenchmarkAnalyzer
from net_benchmark.dns_benchmark.exporters import (
    CSVExporter,
    ExcelExporter,
    ExportBundle,
    PDFExporter,
)


@pytest.fixture
def sample_analyzer(sample_results):
    return BenchmarkAnalyzer(sample_results)


@pytest.fixture
def analyzer(sample_results):
    return BenchmarkAnalyzer(sample_results)


def test_excel_pdf_export(tmp_path, sample_results):
    analyzer = BenchmarkAnalyzer(sample_results)
    excel_path = tmp_path / "report.xlsx"
    pdf_path = tmp_path / "report.pdf"

    # Excel with all sheets
    ExcelExporter.export_results(
        sample_results,
        analyzer,
        str(excel_path),
        domain_stats=analyzer.get_domain_statistics(),
        record_type_stats=analyzer.get_record_type_statistics(),
        error_stats=analyzer.get_error_statistics(),
    )
    assert excel_path.exists() and excel_path.stat().st_size > 0

    # PDF only if weasyprint is available
    if importlib.util.find_spec("weasyprint"):
        PDFExporter.export_results(
            sample_results, analyzer, str(pdf_path), include_success_chart=True
        )
        assert pdf_path.exists() and pdf_path.stat().st_size > 0
    else:
        pytest.skip("weasyprint not installed; skipping PDF export test")


def test_generate_latency_chart_for_excel_creates_file(sample_analyzer):
    tmpdir = tempfile.mkdtemp()
    chart_path = ExcelExporter._generate_latency_chart_for_excel(
        sample_analyzer, tmpdir
    )
    assert os.path.exists(chart_path)
    assert chart_path.endswith(".png")
    assert os.path.getsize(chart_path) > 0


def test_generate_success_chart_for_excel_creates_file(sample_analyzer):
    tmpdir = tempfile.mkdtemp()
    chart_path = ExcelExporter._generate_success_chart_for_excel(
        sample_analyzer, tmpdir
    )
    assert os.path.exists(chart_path)
    assert chart_path.endswith(".png")
    assert os.path.getsize(chart_path) > 0


def test_add_charts_sheet_embeds_images(sample_analyzer):
    wb = Workbook()
    wb.remove(wb.active)
    tmpdir = tempfile.mkdtemp()
    chart_paths = ExcelExporter._add_charts_sheet(wb, sample_analyzer, tmpdir)

    # Verify sheet exists
    assert "Charts" in wb.sheetnames
    ws = wb["Charts"]

    # Check title cell
    assert ws["A1"].value == "DNS Resolver Performance Charts"

    # Ensure chart files were created
    for path in chart_paths:
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0


def test_export_json_creates_valid_file(tmp_path, sample_results, analyzer):
    out_path = tmp_path / "report.json"
    ExportBundle.export_json(
        sample_results,
        analyzer,
        domain_stats=analyzer.get_domain_statistics(),
        record_type_stats=analyzer.get_record_type_statistics(),
        error_stats=analyzer.get_error_statistics(),
        output_path=str(out_path),
    )
    assert out_path.exists()
    data = json.loads(out_path.read_text())
    # Basic structure checks
    assert "overall" in data
    assert "resolver_stats" in data
    assert isinstance(data["raw_results"], list)
    assert "domain_stats" in data
    assert "record_type_stats" in data
    assert "error_stats" in data


def test_export_raw_results_csv(tmp_path, sample_results):
    out_path = tmp_path / "raw.csv"
    CSVExporter.export_raw_results(sample_results, str(out_path))
    df = pd.read_csv(out_path)
    # Ensure expected columns exist
    expected_cols = {
        "timestamp",
        "resolver_name",
        "resolver_ip",
        "domain",
        "record_type",
        "latency_ms",
        "status",
        "answers_count",
        "ttl",
        "error_message",
        "cache_hit",
        "iteration",
        "query_id",
    }
    assert expected_cols.issubset(df.columns)
    assert len(df) == len(sample_results)


def test_export_summary_statistics_csv(tmp_path, analyzer):
    out_path = tmp_path / "summary.csv"
    CSVExporter.export_summary_statistics(analyzer, str(out_path))
    df = pd.read_csv(out_path)
    assert "resolver_name" in df.columns
    assert "avg_latency_ms" in df.columns
    assert len(df) == len(analyzer.get_resolver_statistics())


def test_export_domain_statistics_csv(tmp_path, analyzer):
    out_path = tmp_path / "domains.csv"
    stats = analyzer.get_domain_statistics()
    CSVExporter.export_domain_statistics(stats, str(out_path))
    df = pd.read_csv(out_path)
    assert not df.empty
    assert set(stats[0].keys()).issubset(df.columns)


def test_export_record_type_statistics_csv(tmp_path, analyzer):
    out_path = tmp_path / "record_types.csv"
    stats = analyzer.get_record_type_statistics()
    CSVExporter.export_record_type_statistics(stats, str(out_path))
    df = pd.read_csv(out_path)
    assert not df.empty
    assert set(stats[0].keys()).issubset(df.columns)


def test_export_error_statistics_csv(tmp_path, analyzer):
    out_path = tmp_path / "errors.csv"
    stats = analyzer.get_error_statistics()
    CSVExporter.export_error_statistics(stats, str(out_path))
    df = pd.read_csv(out_path)
    assert "error_message" in df.columns
    assert "count" in df.columns
    assert len(df) == len(stats)


@pytest.mark.skipif(
    importlib.util.find_spec("weasyprint"),
    reason="weasyprint installed; skip missing-dep test",
)
def test_pdf_exporter_raises_without_weasyprint(
    sample_results, analyzer, tmp_path, monkeypatch
):
    # Force module-level HTML to None to simulate missing weasyprint
    monkeypatch.setattr(exporters, "HTML", None)
    pdf_path = tmp_path / "report.pdf"

    with pytest.raises(RuntimeError) as excinfo:
        exporters.PDFExporter.export_results(sample_results, analyzer, str(pdf_path))

    assert "PDF export requires 'weasyprint'" in str(excinfo.value)
