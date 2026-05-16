"""Shared exporter base utilities."""

from typing import Any, List, Tuple

import matplotlib
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

matplotlib.use("Agg")

# ---------------------------------------------------------------------------
# Colour constants reused across all three modules
# ---------------------------------------------------------------------------

FILL_HEADER = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_AMBER = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Bar chart colour thresholds — callers pass their own boundary tuples when the
# defaults don't apply (e.g. SSL uses days_remaining, not latency).
# Default: green < 50 ms, amber < 100 ms, red >= 100 ms.
DEFAULT_BAR_THRESHOLDS: Tuple[float, float] = (50.0, 100.0)


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------


def autosize_columns(ws: Any, max_width: int = 50) -> None:
    """Set each column width to the longest cell value, capped at max_width."""
    for column in ws.columns:
        max_length = max(
            (len(str(cell.value)) for cell in column if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[column[0].column_letter].width = min(
            max_length + 2, max_width
        )


def add_simple_table_sheet(wb: Workbook, title: str, df: pd.DataFrame) -> None:
    """Create a formatted table sheet from a DataFrame.

    Replaces the four verbatim copies in dns_benchmark/exporters.py.
    """
    ws = wb.create_sheet(title)
    for col_idx, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = FILL_HEADER
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    autosize_columns(ws)


def add_coloured_table_sheet(
    wb: Workbook,
    title: str,
    headers: List[str],
    rows: List[List[Any]],
    # row_fill_fn(row_values) → PatternFill — caller decides colour logic
    row_fill_fn: Any,
) -> None:
    """Create a table sheet where each data row is coloured by a caller function.

    Used for DNSSEC sheet (DNS), Security Headers sheet (HTTP),
    and Expiry Timeline sheet (SSL) — all share the same green/amber/red pattern
    but with different logic deciding which colour applies.
    """
    ws = wb.create_sheet(title)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = FILL_HEADER
    for row_idx, row_values in enumerate(rows, 2):
        fill = row_fill_fn(row_values)
        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
    autosize_columns(ws)


def embed_charts_sheet(
    wb: Workbook,
    title: str,
    chart_entries: List[Tuple[str, str, str]],  # (cell_anchor, heading, image_path)
    sheet_title: str = "",
) -> None:
    """Embed pre-generated chart PNGs into a Charts sheet.

    chart_entries: list of (anchor_cell, heading_cell, image_path)
    e.g. [("A4", "A3", "/tmp/latency.png"), ("A24", "A23", "/tmp/success.png")]
    """
    ws = wb.create_sheet(title)
    if sheet_title:
        ws["A1"] = sheet_title
        ws["A1"].font = Font(bold=True, size=14)
    for heading_cell, anchor_cell, img_path in chart_entries:
        if heading_cell:
            ws[heading_cell].font = Font(bold=True, size=12)
        img = XLImage(img_path)
        img.width = 600
        img.height = 360
        ws.add_image(img, anchor_cell)


# ---------------------------------------------------------------------------
# Chart generation
# ---------------------------------------------------------------------------


def generate_bar_chart(
    names: List[str],
    values: List[float],
    ylabel: str,
    title: str,
    output_path: str,
    thresholds: Tuple[float, float] = DEFAULT_BAR_THRESHOLDS,
    value_fmt: str = "{:.1f}",
    figsize: Tuple[int, int] = (10, 6),
    invert_colours: bool = False,
) -> str:
    """Generate a bar chart PNG and save to output_path. Returns output_path.

    thresholds: (low, high) — values below low are green, below high amber, else red.
    invert_colours: True for metrics where higher is better (e.g. success rate,
                    days_remaining) — green is high, red is low.
    value_fmt: format string for bar labels, e.g. "{:.1f}%" for percentages.
    """
    low, high = thresholds

    def _colour(v: float) -> str:
        if invert_colours:
            return "#2ecc71" if v >= high else "#f39c12" if v >= low else "#e74c3c"
        return "#2ecc71" if v < low else "#f39c12" if v < high else "#e74c3c"

    colours = [_colour(v) for v in values]

    fig, ax = plt.subplots(figsize=figsize)
    if not names:
        ax.text(0.5, 0.5, "No data", ha="center", va="center", fontsize=14)
        ax.axis("off")
    else:
        bars = ax.bar(range(len(names)), values, color=colours)
        ax.set_xticks(range(len(names)))
        ax.set_xticklabels(names, rotation=45, ha="right")
        ax.set_ylabel(ylabel)
        ax.set_title(title)
        for bar in bars:
            h = bar.get_height()
            ax.annotate(
                value_fmt.format(h),
                xy=(bar.get_x() + bar.get_width() / 2, h),
                xytext=(0, 3),
                textcoords="offset points",
                ha="center",
                va="bottom",
                fontsize=8,
            )
        ax.grid(True, alpha=0.3, axis="y")

    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close()
    return output_path


# ---------------------------------------------------------------------------
# PDF / HTML shared CSS
# ---------------------------------------------------------------------------

BASE_CSS = """
    body {
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        margin: 0; padding: 20px; color: #333; line-height: 1.6;
    }
    .header {
        text-align: center;
        border-bottom: 3px solid #2c3e50;
        padding-bottom: 20px;
        margin-bottom: 30px;
    }
    .section { margin-bottom: 40px; }
    table {
        width: 100%; border-collapse: collapse;
        margin: 20px 0; font-size: 0.9em;
    }
    th, td { padding: 10px 14px; text-align: left; border-bottom: 1px solid #ddd; }
    th { background-color: #34495e; color: white; }
    tr:nth-child(even) { background-color: #f8f9fa; }
    .badge-ok      { color: #27ae60; font-weight: bold; }
    .badge-warn    { color: #e67e22; font-weight: bold; }
    .badge-crit    { color: #e74c3c; font-weight: bold; }
    .badge-missing { color: #95a5a6; }
    .chart { text-align: center; margin: 30px 0; }
    .chart img {
        max-width: 100%; height: auto;
        border: 1px solid #ddd; border-radius: 5px;
    }
    .alert-box {
        background: #ffeeba; border: 1px solid #ffc107;
        border-radius: 4px; padding: 12px 16px; margin: 10px 0;
    }
"""


def html_page(title: str, body: str) -> str:
    """Wrap body HTML in a full page with shared CSS."""
    return f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="UTF-8">
  <title>{title}</title>
  <style>{BASE_CSS}</style>
</head>
<body>
{body}
</body>
</html>"""
