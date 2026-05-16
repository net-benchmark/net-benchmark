"""Export functionality for HTTP benchmark results."""

import base64
import json
import os
import tempfile
from datetime import datetime
from typing import Dict, List, Optional

import pandas as pd

try:
    from weasyprint import HTML
except ImportError:
    HTML = None

from openpyxl import Workbook
from openpyxl.styles import Font

from net_benchmark.dns_benchmark.core import QueryStatus
from net_benchmark.exporters.base import (
    FILL_HEADER,
    add_simple_table_sheet,
    autosize_columns,
    embed_charts_sheet,
    generate_bar_chart,
    html_page,
)
from net_benchmark.http_bench.analysis import HTTPAnalyzer
from net_benchmark.http_bench.core import SECURITY_HEADERS, HTTPResult

# ---------------------------------------------------------------------------
# JSON bundle
# ---------------------------------------------------------------------------


class HTTPExportBundle:
    @staticmethod
    def export_json(
        results: List[HTTPResult],
        analyzer: HTTPAnalyzer,
        output_path: str,
    ) -> None:
        payload = {
            "overall": analyzer.get_overall_statistics(),
            "target_stats": [vars(s) for s in analyzer.get_target_statistics()],
            "protocol_distribution": analyzer.get_protocol_distribution(),
            "ttfb_statistics": analyzer.get_ttfb_statistics(),
            "security_summary": analyzer.get_security_summary(),
            "status_code_distribution": analyzer.get_status_code_distribution(),
            "error_stats": analyzer.get_error_statistics(),
            "raw_results": [r.to_dict() for r in results],
        }
        with open(output_path, "w") as f:
            json.dump(payload, f, indent=2, default=str)


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


class HTTPCSVExporter:
    """Mirrors dns_benchmark CSVExporter — one static method per export type."""

    @staticmethod
    def export_raw_results(results: List[HTTPResult], output_path: str) -> None:
        df = pd.DataFrame([r.to_dict() for r in results])
        df.to_csv(output_path, index=False)

    @staticmethod
    def export_summary_statistics(analyzer: HTTPAnalyzer, output_path: str) -> None:
        data = []
        for s in analyzer.get_target_statistics():
            data.append(
                {
                    "target": s.target,
                    "method": s.method,
                    "total_requests": s.total_requests,
                    "successful_requests": s.successful_requests,
                    "success_rate": s.success_rate,
                    "min_latency_ms": s.min_latency,
                    "avg_latency_ms": s.avg_latency,
                    "median_latency_ms": s.median_latency,
                    "max_latency_ms": s.max_latency,
                    "p95_latency_ms": s.p95_latency,
                    "p99_latency_ms": s.p99_latency,
                    "jitter_ms": s.jitter,
                    "consistency_score": s.consistency_score,
                    "avg_ttfb_ms": s.avg_ttfb_ms,
                    "p95_ttfb_ms": s.p95_ttfb_ms,
                    "http2_rate": s.http2_rate,
                    "redirect_rate": round(s.redirect_rate, 2),
                    "avg_response_size_bytes": round(s.avg_response_size_bytes, 2),
                    "avg_dns_ms": round(s.avg_dns_ms, 2),
                    "avg_tcp_ms": round(s.avg_tcp_ms, 2),
                    "avg_tls_ms": round(s.avg_tls_ms, 2),
                    "avg_compressed_size_bytes": round(s.avg_compressed_size_bytes, 2),
                    "avg_redirect_time_ms": round(s.avg_redirect_time_ms, 2),
                    "http2_downgrade_rate": round(s.http2_downgrade_rate, 2),
                    "cdn_fingerprint": s.cdn_fingerprint or "",
                    "server_header": s.server_header or "",
                    "cert_expiry_days_min": s.cert_expiry_days_min,
                    "cache_control_present": s.cache_control_present,
                    "etag_present": s.etag_present,
                    "last_modified_present": s.last_modified_present,
                    "age_present": s.age_present,
                }
            )
        pd.DataFrame(data).to_csv(output_path, index=False)

    @staticmethod
    def export_security_statistics(analyzer: HTTPAnalyzer, output_path: str) -> None:
        """Per‑target security header presence matrix."""
        # Build presence matrix
        matrix = []
        for target_stat in analyzer.get_target_statistics():
            target = target_stat.target
            row = {"target": target}
            # Get all raw results for this target (successful only)
            for h in SECURITY_HEADERS:
                present = False
                for r in analyzer.results:
                    if r.target == target and r.status == QueryStatus.SUCCESS:
                        if r.security_headers.get(h) is not None:
                            present = True
                            break
                row[h] = "✓" if present else "✗"
            matrix.append(row)
        pd.DataFrame(matrix).to_csv(output_path, index=False)

    @staticmethod
    def export_ttfb_statistics(analyzer: HTTPAnalyzer, output_path: str) -> None:
        pd.DataFrame(analyzer.get_ttfb_statistics()).to_csv(output_path, index=False)

    @staticmethod
    def export_protocol_statistics(analyzer: HTTPAnalyzer, output_path: str) -> None:
        pd.DataFrame(analyzer.get_protocol_distribution()).to_csv(
            output_path, index=False
        )

    @staticmethod
    def export_error_statistics(analyzer: HTTPAnalyzer, output_path: str) -> None:
        errors = analyzer.get_error_statistics()
        pd.DataFrame(
            [{"error_message": k, "count": v} for k, v in errors.items()]
        ).to_csv(output_path, index=False)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


class HTTPExcelExporter:
    """Mirrors dns_benchmark ExcelExporter.
    Uses base helpers — no duplicated sheet/chart code.
    """

    @staticmethod
    def export_results(
        results: List[HTTPResult],
        analyzer: HTTPAnalyzer,
        output_path: str,
        error_stats: Optional[Dict[str, int]] = None,
        include_charts: bool = False,
    ) -> None:
        wb = Workbook()
        wb.remove(wb.active)

        temp_dir = None
        chart_paths: List[str] = []

        try:
            HTTPExcelExporter._add_raw_data_sheet(wb, results)
            HTTPExcelExporter._add_target_summary_sheet(wb, analyzer)
            HTTPExcelExporter._add_ttfb_sheet(wb, analyzer)
            HTTPExcelExporter._add_security_headers_sheet(wb, analyzer)

            proto_stats = analyzer.get_protocol_distribution()
            if proto_stats:
                add_simple_table_sheet(wb, "Protocol", pd.DataFrame(proto_stats))

            status_dist = analyzer.get_status_code_distribution()
            if status_dist:
                add_simple_table_sheet(wb, "Status Codes", pd.DataFrame(status_dist))

            if error_stats:
                add_simple_table_sheet(
                    wb,
                    "Errors",
                    pd.DataFrame(
                        [{"error": k, "count": v} for k, v in error_stats.items()]
                    ),
                )

            if include_charts:
                temp_dir = tempfile.mkdtemp()
                chart_paths = HTTPExcelExporter._add_charts_sheet(
                    wb, analyzer, temp_dir
                )

            wb.save(output_path)
        finally:
            for p in chart_paths:
                try:
                    if os.path.exists(p):
                        os.remove(p)
                except OSError:
                    pass
            if temp_dir and os.path.exists(temp_dir):
                try:
                    os.rmdir(temp_dir)
                except OSError:
                    pass

    @staticmethod
    def _add_raw_data_sheet(wb: Workbook, results: List[HTTPResult]) -> None:
        data = []
        for r in results:
            present_headers = "|".join(
                h for h, v in r.security_headers.items() if v is not None
            )
            data.append(
                {
                    "Target": r.target,
                    "Method": r.method,
                    "Status": r.status.value,
                    "HTTP Code": r.http_status_code or "",
                    "Total (ms)": round(r.total_ms, 2),
                    "TTFB (ms)": round(r.ttfb_ms, 2) if r.ttfb_ms else "",
                    "Protocol": r.protocol.value,
                    "Redirects": r.redirect_count,
                    "Size (bytes)": r.response_size_bytes or "",
                    "Compressed": r.compressed,
                    "Content-Type": r.content_type or "",
                    "CDN": r.cdn_fingerprint or "",
                    "Server": r.server_header or "",
                    "Cert Expiry (days)": (
                        r.cert_expiry_days if r.cert_expiry_days is not None else ""
                    ),
                    "Alt-Svc": r.alt_svc or "",
                    "IP Version": r.ip_version or "",
                    "Security Headers": present_headers,
                    "Iteration": r.iteration,
                    "Attempt": r.attempt_number,
                    "Error": r.error_message or "",
                }
            )
        add_simple_table_sheet(wb, "Raw Data", pd.DataFrame(data))

    @staticmethod
    def _add_target_summary_sheet(wb: Workbook, analyzer: HTTPAnalyzer) -> None:
        data = []
        for s in analyzer.get_target_statistics():
            data.append(
                {
                    "Target": s.target,
                    "Method": s.method,
                    "Total": s.total_requests,
                    "Successful": s.successful_requests,
                    "Success Rate (%)": round(s.success_rate, 2),
                    "Avg (ms)": round(s.avg_latency, 2),
                    "Median (ms)": round(s.median_latency, 2),
                    "P95 (ms)": round(s.p95_latency, 2),
                    "P99 (ms)": round(s.p99_latency, 2),
                    "Jitter": round(s.jitter, 2),
                    "Consistency": round(s.consistency_score, 2),
                    "Avg TTFB (ms)": round(s.avg_ttfb_ms, 2),
                    "HTTP/2 Rate (%)": round(s.http2_rate, 2),
                    "Redirect Rate (%)": round(s.redirect_rate, 2),
                    "Avg Size (bytes)": round(s.avg_response_size_bytes, 2),
                    "Avg DNS (ms)": round(s.avg_dns_ms, 2),
                    "Avg TCP (ms)": round(s.avg_tcp_ms, 2),
                    "Avg TLS (ms)": round(s.avg_tls_ms, 2),
                    "Avg Compressed Size (bytes)": round(
                        s.avg_compressed_size_bytes, 2
                    ),
                    "Avg Redirect Time (ms)": round(s.avg_redirect_time_ms, 2),
                    "HTTP/2 Downgrade Rate (%)": round(s.http2_downgrade_rate, 2),
                    "Cache-Control": s.cache_control_present,
                    "ETag": s.etag_present,
                    "Last-Modified": s.last_modified_present,
                    "Age": s.age_present,
                    "CDN": s.cdn_fingerprint or "",
                    "Server": s.server_header or "",
                    "Cert Expiry (days)": (
                        s.cert_expiry_days_min
                        if s.cert_expiry_days_min is not None
                        else ""
                    ),
                    "Alt-Svc": s.alt_svc or "",
                    "IP Version": s.ip_version or "",
                }
            )
        add_simple_table_sheet(wb, "Target Summary", pd.DataFrame(data))

    @staticmethod
    def _add_ttfb_sheet(wb: Workbook, analyzer: HTTPAnalyzer) -> None:
        add_simple_table_sheet(
            wb, "TTFB Analysis", pd.DataFrame(analyzer.get_ttfb_statistics())
        )

    @staticmethod
    def _add_security_headers_sheet(wb: Workbook, analyzer: HTTPAnalyzer) -> None:
        """Security headers presence sheet — one row per target, each header cell
        coloured green (✓) or red (✗). Target column has no fill.
        """
        from openpyxl.styles import PatternFill

        # Build per‑target presence from successful results only
        per_target: Dict[str, Dict[str, bool]] = {}
        for r in analyzer.results:
            if r.status != QueryStatus.SUCCESS:
                continue
            if r.target not in per_target:
                per_target[r.target] = {}
            for h in SECURITY_HEADERS:
                if r.security_headers.get(h) is not None:
                    per_target[r.target][h] = True

        headers_cols = ["Target"] + [
            h.title().replace("-", " ") for h in SECURITY_HEADERS
        ]
        rows = []
        for target_stat in analyzer.get_target_statistics():
            target = target_stat.target
            presence = per_target.get(target, {})
            row = [target] + ["✓" if presence.get(h) else "✗" for h in SECURITY_HEADERS]
            rows.append(row)

        ws = wb.create_sheet("Security Headers")

        # Write header row
        for col_idx, header in enumerate(headers_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = FILL_HEADER

        # Write data rows with per‑cell colours
        green_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )

        for row_idx, row_values in enumerate(rows, 2):
            for col_idx, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 1:  # Target column – no fill
                    continue
                cell.fill = green_fill if value == "✓" else red_fill

        autosize_columns(ws)

    @staticmethod
    def _add_charts_sheet(
        wb: Workbook, analyzer: HTTPAnalyzer, temp_dir: str
    ) -> List[str]:
        target_stats = analyzer.get_target_statistics()
        valid = [s for s in target_stats if s.successful_requests > 0]

        names = [s.target.replace("https://", "").replace("http://", "") for s in valid]

        latency_path = generate_bar_chart(
            names=names,
            values=[s.avg_latency for s in valid],
            ylabel="Avg Latency (ms)",
            title="HTTP Target Latency Comparison",
            output_path=os.path.join(temp_dir, "http_latency.png"),
        )
        ttfb_path = generate_bar_chart(
            names=names,
            values=[s.avg_ttfb_ms for s in valid],
            ylabel="Avg TTFB (ms)",
            title="Time to First Byte Comparison",
            output_path=os.path.join(temp_dir, "http_ttfb.png"),
        )
        success_path = generate_bar_chart(
            names=names,
            values=[s.success_rate for s in valid],
            ylabel="Success Rate (%)",
            title="HTTP Target Success Rates",
            output_path=os.path.join(temp_dir, "http_success.png"),
            thresholds=(80.0, 95.0),
            invert_colours=True,
            value_fmt="{:.1f}%",
        )

        embed_charts_sheet(
            wb,
            "Charts",
            [
                ("A3", "A4", latency_path),
                ("A23", "A24", ttfb_path),
                ("A43", "A44", success_path),
            ],
            "HTTP Benchmark Charts",
        )
        return [latency_path, ttfb_path, success_path]


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


class HTTPPDFExporter:
    """Mirrors dns_benchmark PDFExporter."""

    @staticmethod
    def export_results(
        results: List[HTTPResult],
        analyzer: HTTPAnalyzer,
        output_path: str,
        include_charts: bool = True,
    ) -> None:
        if HTML is None:
            raise RuntimeError(
                "PDF export requires 'weasyprint'. "
                "Install with: pip install net-benchmark[pdf]"
            )

        charts_dir = tempfile.mkdtemp()
        chart_paths: List[str] = []

        try:
            target_stats = analyzer.get_target_statistics()
            valid = [s for s in target_stats if s.successful_requests > 0]
            names = [
                s.target.replace("https://", "").replace("http://", "") for s in valid
            ]

            # Generate all three charts
            latency_path = generate_bar_chart(
                names=names,
                values=[s.avg_latency for s in valid],
                ylabel="Avg Latency (ms)",
                title="HTTP Target Latency Comparison",
                output_path=os.path.join(charts_dir, "latency.png"),
            )
            ttfb_path = generate_bar_chart(
                names=names,
                values=[s.avg_ttfb_ms for s in valid],
                ylabel="Avg TTFB (ms)",
                title="Time to First Byte Comparison",
                output_path=os.path.join(charts_dir, "ttfb.png"),
            )
            success_path = generate_bar_chart(
                names=names,
                values=[s.success_rate for s in valid],
                ylabel="Success Rate (%)",
                title="HTTP Target Success Rates",
                output_path=os.path.join(charts_dir, "success.png"),
                thresholds=(80.0, 95.0),
                invert_colours=True,
                value_fmt="{:.1f}%",
            )
            chart_paths.extend([latency_path, ttfb_path, success_path])

            # Read images as base64
            with open(latency_path, "rb") as f:
                latency_b64 = base64.b64encode(f.read()).decode()
            with open(ttfb_path, "rb") as f:
                ttfb_b64 = base64.b64encode(f.read()).decode()
            with open(success_path, "rb") as f:
                success_b64 = base64.b64encode(f.read()).decode()

            html_content = HTTPPDFExporter._generate_html(
                analyzer, latency_b64, ttfb_b64, success_b64
            )
            HTML(string=html_content).write_pdf(output_path)

        finally:
            for p in chart_paths:
                try:
                    if os.path.exists(p):
                        os.remove(p)
                except OSError:
                    pass
            try:
                os.rmdir(charts_dir)
            except OSError:
                pass

    @staticmethod
    def _generate_html(
        analyzer: HTTPAnalyzer,
        latency_b64: str,
        ttfb_b64: str,
        success_b64: str,
    ) -> str:
        overall = analyzer.get_overall_statistics()
        target_stats = analyzer.get_target_statistics()
        security = analyzer.get_security_summary()
        ranked = sorted(
            [s for s in target_stats if s.successful_requests > 0],
            key=lambda s: s.avg_latency,
        )
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # --- target rankings table ---
        ranking_rows = "".join(
            f"<tr><td>{i + 1}</td><td>{s.target}</td>"
            f"<td>{s.avg_latency:.1f}</td><td>{s.avg_ttfb_ms:.1f}</td>"
            f"<td>{s.success_rate:.1f}%</td>"
            f"<td>{'HTTP/2' if s.http2_rate > 50 else 'HTTP/1.1'}</td>"
            f"<td>{s.ip_version or 'N/A'}</td>"
            f"<td>{'✓' if s.alt_svc else '✗'}</td></tr>"
            for i, s in enumerate(ranked)
        )

        # --- security summary table ---
        header_counts = security.get("security_header_counts", {})
        total_req = security.get("total_requests", 1)
        sec_rows = "".join(
            f"<tr><td>{h.title()}</td>"
            f"<td>{header_counts.get(h, 0)}</td>"
            f"<td>{header_counts.get(h, 0) / total_req * 100:.1f}%</td>"
            f"<td class='{'badge-ok' if header_counts.get(h, 0) > 0 else 'badge-crit'}'>"
            f"{'✓' if header_counts.get(h, 0) > 0 else '✗'}</td>"
            f"</tr>"
            for h in SECURITY_HEADERS
        )

        cdn_dist = security.get("cdn_distribution", {})
        cdn_str = (
            ", ".join(f"{k} ({v})" for k, v in cdn_dist.items()) or "None detected"
        )

        # Compute real minimum certificate expiry from target stats
        min_expiry = None
        for s in target_stats:
            if s.cert_expiry_days_min is not None:
                if min_expiry is None or s.cert_expiry_days_min < min_expiry:
                    min_expiry = s.cert_expiry_days_min
        cert_expiry_line = f"<p><strong>Cert expiry (worst):</strong> {min_expiry if min_expiry is not None else 'N/A'} days</p>"

        # Collect alt_svc and ip_version signals from target stats
        alt_svc_targets = [s for s in target_stats if s.alt_svc]
        alt_svc_str = (
            ", ".join(
                f"{s.target.replace('https://', '').replace('http://', '')}: {s.alt_svc}"
                for s in alt_svc_targets
            )
            or "None detected"
        )

        ip_versions = set(s.ip_version for s in target_stats if s.ip_version)
        ip_version_str = ", ".join(sorted(ip_versions)) or "N/A"

        body = f"""
        <div class="header">
        <h1>HTTP Benchmark Report</h1>
        <p>Generated: {now}</p>
        </div>

        <div class="section">
        <h2>Executive Summary</h2>
        <p><strong>Total requests:</strong> {overall['total_requests']}</p>
        <p><strong>Successful:</strong> {overall['successful_requests']} ({overall['overall_success_rate']:.1f}%)</p>
        <p><strong>Avg latency:</strong> {overall['overall_avg_latency']:.1f} ms</p>
        <p><strong>Avg TTFB:</strong> {overall['overall_avg_ttfb']:.1f} ms</p>
        <p><strong>HTTP/2 rate:</strong> {overall['http2_rate']:.1f}%</p>
        <p><strong>HSTS coverage:</strong> {overall['hsts_coverage']:.1f}%</p>
        <p><strong>Targets tested:</strong> {overall['target_count']}</p>
        <p><strong>Fastest target:</strong> {overall['fastest_target']}</p>
        <p><strong>Slowest target:</strong> {overall['slowest_target']}</p>
        {cert_expiry_line}
        <p><strong>IP version(s):</strong> {ip_version_str}</p>
        <p><strong>HTTP/3 advertised (Alt-Svc):</strong> {alt_svc_str}</p>
        </div>

        <div class="section">
        <h2>Latency Comparison</h2>
        <div class="chart"><img src="data:image/png;base64,{latency_b64}" alt="Latency Comparison"></div>
        </div>

        <div class="section">
        <h2>Time to First Byte (TTFB)</h2>
        <div class="chart"><img src="data:image/png;base64,{ttfb_b64}" alt="TTFB Comparison"></div>
        </div>

        <div class="section">
        <h2>Success Rates</h2>
        <div class="chart"><img src="data:image/png;base64,{success_b64}" alt="Success Rate Comparison"></div>
        </div>

        <div class="section">
        <h2>Target Rankings</h2>
        <table>
        <tr><th>Rank</th><th>Target</th><th>Avg (ms)</th><th>TTFB (ms)</th><th>Success</th><th>Protocol</th><th>IP Ver</th><th>HTTP/3</th></tr>
            {ranking_rows}
        </table>
        </div>

        <div class="section">
        <h2>Security Headers Audit</h2>
        <p><strong>CDN detected:</strong> {cdn_str}</p>
        <p><strong>Server header leaks:</strong> {security.get('server_header_leak_count', 0)}</p>
        <table>
            <tr><th>Header</th><th>Present Count</th><th>Coverage</th><th>Status</th></tr>
            {sec_rows}
        </table>
        </div>
        """
        return html_page("HTTP Benchmark Report", body)
